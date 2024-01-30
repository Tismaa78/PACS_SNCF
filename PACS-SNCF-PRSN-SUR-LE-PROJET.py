# début du code
import pandas as pd
from tkinter import *
from PIL import Image, ImageTk
from datetime import *
import time
from tkinter import messagebox
from tkinter import filedialog
import sys
from tkinter import ttk
import openpyxl
from datetime import datetime
from tkinter import messagebox, Tk, Label
from tkinter import messagebox, Tk
import docx
from docx import Document
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
from openpyxl import load_workbook


class Dashboard:
    def __init__(self, root):
        self.root = root
        self.root.title("SNCF PACS")
        self.root.geometry("1366x768")
        self.root.config(bg="#eff5f6")

        icon = PhotoImage(file=r"C:\Users\mathi\Desktop\Pacs SNCF\Image\pic-icon.png")
        self.root.iconphoto(True, icon)

        ### En-tête
        self.entete = Frame(self.root, bg="#009df4")
        self.entete.place(x=300, y=0, width=1070, height=60)

        self.deconnecte = Button(self.entete, text="Deconnecter", bg="#FF0000", font=("times new roman", 13,"bold"), bd=0, fg="white", cursor="hand2", activebackground="#32cf8e", command=self.deconnecter_application)
        self.deconnecte.place(x=950, y=15)

        # Menu
        self.FrameMenu = Frame(self.root, bg="#ffffff")
        self.FrameMenu.place(x=0, y=0, width=300, height=750)

        self.logoImage = Image.open(r"C:\Users\mathi\Desktop\Pacs SNCF\Image\kontron2.png")
        photo = ImageTk.PhotoImage(self.logoImage)
        self.logo = Label(self.FrameMenu, image=photo, bg="#ffffff")
        self.logo.Image = photo
        self.logo.place(x=20, y=50)

        self.Nom = Label(self.FrameMenu, text="PACS SNCF", bg="#ffffff", font=("times new roman", 13, "bold"))
        self.Nom.place(x=80, y=150)

        # Tableau De Bord
        self.dashboardImage = Image.open(r"C:\Users\mathi\Desktop\Pacs SNCF\Image\dashboard-icon.png")
        photo = ImageTk.PhotoImage(self.dashboardImage)
        self.dashboard = Label(self.FrameMenu, image=photo, bg="#ffffff")
        self.dashboard.Image = photo
        self.dashboard.place(x=35, y=289)

        self.dashboard_text = Button(self.FrameMenu, text="Tableau de Bord", bg="#ffffff",font=("times new roman", 13, "bold"), bd=0, cursor="hand2", activebackground="#ffffff")
        self.dashboard_text.place(x=80, y=289)

        # Gestion
        self.gestionImage = Image.open(r"C:\Users\mathi\Desktop\Pacs SNCF\Image\manage-icon.png")
        photo = ImageTk.PhotoImage(self.gestionImage)
        self.gestion = Label(self.FrameMenu, image=photo, bg="#ffffff")
        self.gestion.Image = photo
        self.gestion.place(x=35, y=340)

        self.gestion_text = Button(self.FrameMenu, text="Gestion", bg="#ffffff",font=("times new roman", 13, "bold"), bd=0, cursor="hand2", activebackground="#ffffff")
        self.gestion_text.place(x=80, y=345)

        # Parametre
        self.parametreImage = Image.open(r"C:\Users\mathi\Desktop\Pacs SNCF\Image\settings-icon.png")
        photo = ImageTk.PhotoImage(self.parametreImage)
        self.parametre = Label(self.FrameMenu, image=photo, bg="#ffffff")
        self.parametre.Image = photo
        self.parametre.place(x=35, y=402)
        
        self.parametre_text = Button(self.FrameMenu, text="Paramètres", bg="#ffffff",font=("times new roman", 13, "bold"), bd=0, cursor="hand2", activebackground="#ffffff")
        self.parametre_text.place(x=80, y=402)

        # Quitter
        self.quitterImage = Image.open(r"C:\Users\mathi\Desktop\Pacs SNCF\Image\exit-icon.png")
        photo = ImageTk.PhotoImage(self.quitterImage)
        self.quitter = Label(self.FrameMenu, image=photo, bg="#ffffff")
        self.quitter.Image = photo
        self.quitter.place(x=25, y=452)

        self.quitter_text = Button(self.FrameMenu, text="Quitter", bg="#ffffff",font=("times new roman", 13, "bold"), bd=0, cursor="hand2", activebackground="#ffffff", command=self.quitter_application)
        self.quitter_text.place(x=85, y=462)

        # corps
        self.titre = Label(self.root, text="Tableau de Bord", font=("times new roman", 13, "bold"), fg="#0064d3", bg="#eff5f6")
        self.titre.place(x=325, y=70)

        # Corp1
        self.corp1 = Frame(self.root, bg="#ffffff")
        self.corp1.place(x=328, y=110, width=1010, height=200)

        # Corp2
        self.corp2 = Frame(self.root, bg="#ffffff")
        self.corp2.place(x=328, y=350, width=1010, height=200)

        # Bouton d'analyse
        self.analyser_btn = Button(self.corp1, text="Analyser MAJ malware", bg="#32cf8e", font=("times new roman", 13,"bold"), bd=0, fg="white", cursor="hand2", activebackground="#32cf8e", command=self.analyser_fichier)
        self.analyser_btn.place(x=10, y=10)
        
        
        # attribut 
        self.selected_file2 = None

        # Bouton "Analyser version Checkpoint" pour corp2
        self.analyser_checkpoint_btn = Button(self.corp2, text="Analyser version Endpoint", bg="#32cf8e", font=("times new roman", 13, "bold"), bd=0, fg="white", cursor="hand2", activebackground="#32cf8e", command=self.analyser_checkpoint_fichier)
        self.analyser_checkpoint_btn.place(x=10, y=10)
        
        # Ajout du carré pour faire glisser le fichier Excel
        self.drop_zone = Label(self.corp1, text="Sélectionnez le fichier Excel ici", bg="#ffffff", font=("times new roman", 12), bd=1, relief="solid", padx=10, pady=10)
        self.drop_zone.place(x=330, y=10, width=400, height=80)
        self.drop_zone.bind("<Button-1>", self.select_file)
        self.drop_zone.bind("<B1-Motion>", self.drag_file)
        self.drop_zone.bind("<ButtonRelease-1>", self.release_file)
        self.drop_zone.bind("<Enter>", self.on_enter)
        self.drop_zone.bind("<Leave>", self.on_leave)

        self.dragging = False

        # Zone de dépôt du fichier Excel pour corp2
        self.drop_zone2 = Label(self.corp2, text="Sélectionnez le fichier Excel ici", bg="#ffffff", font=("times new roman", 12), bd=1, relief="solid", padx=10, pady=10)
        self.drop_zone2.place(x=330, y=10, width=400, height=80)
        self.drop_zone2.bind("<Button-1>", self.select_file2)
        self.drop_zone2.bind("<B1-Motion>", self.drag_file2)
        self.drop_zone2.bind("<ButtonRelease-1>", self.release_file2)
        self.drop_zone2.bind("<Enter>", self.on_enter2)
        self.drop_zone2.bind("<Leave>", self.on_leave2)

        # Ajouter les attributs des labels
        self.label_nb_personnes_moins_de_48h = Label(self.root, text="")
        self.label_nb_personnes_moins_de_96h = Label(self.root, text="")
        self.label_nb_personnes_moins_de_168h = Label(self.root, text="")
        self.label_nb_personnes_plus_de_168h = Label(self.root, text="")


    def select_file(self, event):
        filename = filedialog.askopenfilename(initialdir="/", title="Select file", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
        self.drop_zone.config(text=filename)

    def drag_file(self, event):
        self.drop_zone.config(bg="#32cf8e")
        self.dragging = True

    def release_file(self, event):
        self.drop_zone.config(bg="#ffffff")
        self.dragging = False

    def on_enter(self, event):
        if self.dragging:
            self.drop_zone.config(bg="#32cf8e")

    def on_leave(self, event):
        if self.dragging:
            self.drop_zone.config(bg="#ffffff")

    def analyser_fichier(self):
        # Vérifier si un fichier a été sélectionné
        filename = self.drop_zone.cget("text")
        if not filename:
            messagebox.showwarning("Avertissement", "Veuillez sélectionner un fichier Excel.")
            return

        # Charger le fichier Excel
        try:
            df = pd.read_excel(filename)
            workbook = load_workbook(filename)
            date_creation = workbook.properties.created.date()
        except:
            messagebox.showerror("Erreur", "Impossible de charger le fichier Excel.")
            return
        
        # Dictionnaire des associations entre les Computer Name et les noms/prénoms
        computer_name_associations = {
            "KTR-FR-NB0037": "Bobi Andrieski",
            "KTR-FR-NB0090": "Daniel Bouilly",
            "KTR-FR-NB0093": "Philippe Dardenne",
            "KTR-FR-NB0094": "Christophe Leroux",
            "KTR-FR-NB0086": "Fernand Mendes",
            "KTR-FR-NB0040": "Cedric Valynseele",
            "KTR-FR-NB0192": "Youssef Tabaddine",
            "KTR-AT-NB0155": "Martin Hollmuller",
            "KTR-AT-NB0153": "Gunther Lempert",
            "KTR-AT-NB0157": "Max Linhart",
            "KTR-FR-NB0116": "Jalel Bouachir",
            "KTR-FR-NB0065": "Victor Cardoso",
            "KTR-FR-NB0110": "Olivier Daunas",
            "KTR-FR-NB0124": "Nicolas De Bouvier",
            "KTR-FR-NB0100": "Driss Erouah",
            "KTR-FR-NB0099": "Jean Paul Fourdrinois",
            "KTR-FR-NB0056": "Pascal Garimede",
            "KTR-FR-NB0115": "Olivier Gouzerh",
            "KTR-FR-NB0024": "Samir Hamiane",
            "KTR-FR-NB0140": "Frederic Jacotot",
            "KTR-FR-NB0117": "Frederic Jan",
            "KTR-FR-NB0097": "Dominique Nabais",
            "KTR-FR-NB0075": "Christine Pietri",
            "KTR-FR-NB0154": "Tsiry Lalaina Ramanantsialonina",
            "KTR-FR-NB0145": "Christophe Surdive",
            "KTR-FR-NB0190": "Olivier Taieb",
            "KTR-FR-NB0121": "Thierry Vincent",
            "KTR-FR-NB0092": "Redouane Dahmani",
            "KTR-AT-NB0125": "Gerald Balham",
            "KTR-DE-NB0029": "Quang-Hai Dao",
            "KTR-AT-NB0242": "György Hepka",
            "KTR-AT-NB0209": "Hüseyin Kahraman",
            "KTR-AT-NB0128": "Robert Klevar",
            "KTR-AT-NB0068": "Thomas Linsberger",
            "KTR-AT-NB0143": "Michael Neumann",
            "KTR-AT-NB0139": "Rezart Para",
            "KTR-AT-NB0223": "Helmut Scheiber",
            "KTR-AT-NB0249": "Hermann Schwarzinger",
            "KTR-AT-NB0164": "Christian Hans Stark",
            "KTR-AT-NB0064": "Christoph Strasser",
            "KTR-AT-NB0219": "Dominik Szanva",
            "KTR-AT-NB0220": "Jeremias Szlezak",
            "KTR-AT-NB0326": "Daniel Urabl",
            "KTR-AT-NB0135": "Eduard Wittmann",
            "KTR-AT-NB0238": "Vladislav Zinovenko",
            "KTR-AT-NB0318": "Luca Zinicola",
            "KTR-AT-NB0319": "Omar Zubair",
            "KTR-PT-NB0209": "Jorge Caetano",
            "KTR-PT-NB0041": "Elizabeth Fernandez",
            "KTR-PT-NB0213": "Ricardo Monteiro",
            "KTR-PT-NB0212": "Bruno Oliveira",
            "KTR-PT-NB0206": "Filipe Soares Pereira",
            "KTR-PT-NB0208": "Filipe Pires",
            "KTR-PT-NB0211": "Miguel Ramos",
            "KTR-PT-NB0201": "Susana Rocha",
            "KTR-PT-NB0210": "Nuno Santos",
            "KTR-PT-NB0214": "Igor Videra Soares",
            "KTR-FR-NB0028": "Kaci Anis Derridi",
            "KTR-FR-NB0029": "Hélène Farine",
            "KTR-FR-NB0135": "Mickael Lambert",
            "KTR-FR-NB0031": "Ludovic Martin",
            "KTR-FR-NB0180": "Mamadou Seck",
            "KTR-FR-NB0035": "Frederic Pesson",
            "KTR-FR-NB0157": "Thibault Rabe",
            "KTR-FR-NB0146": "Elhadi Thierno Sadou Diallo",
            "KTR-FR-NB0130": "Sylia Baraka",
            "KTR-FR-NB0074": "Alain Caille",
            "KTR-FR-NB0041": "Claudy Cavallari",
            "KTR-FR-NB0104": "Jean-Tycho Dollfus",
            "KTR-FR-NB0128": "Laurent Dorys-Charnalet",
            "KTR-FR-NB0060": "David Gastao",
            "KTR-FR-NB0073": "Trung Lim Cao",
            "KTR-FR-NB0166": "Antoine Jegoux",
            "KTR-FR-NB0106": "Olivier Marty",
            "KTR-FR-NB0064": "Benjamina Rabarison",
            "KTR-FR-NB0061": "Florent Roblin",
            "KTR-FR-NB0022": "David Roux",
            "KTR-AT-NB0245": "Levente Csakvari",
            "KTR-FR-NB0127": "Thierry Demarche",
            "KTR-FR-NB0047": "Fatiha Devaud",
            "KTR-FR-NB0132": "Lony Fischer",
            "KTR-FR-NB0138": "Laroussi Gattoufi",
            "KTR-FR-NB0101": "Yann Le Cor",
            "KTR-AT-NB0049": "Andreas Mayrhofer",
            "KTR-AT-NB0312": "Dominik Neznik",
            "KTR-FR-NB0174": "Reva Rabenandrasana",
            "KTR-DE-NB0021": "Ole Hohlt",
            "KTR-DE-NB0042": "Ndubueze Ogaluzo",
            "KTR-AT-NB0241": "Atanas Pechevski",
            "KTR-AT-NB0045": "Martin Riegler",
            "KTR-BE-NB0093": "Endri Stefani",
            "KTR-DE-NB0009": "Erich Seitz",
            "KTR-UK-NB0020": "Therese Taylor",
            "KTR-DE-NB0040": "Onur Ulu",
            "KTR-AT-NB0156": "Gunter Bauer",
            "KTR-AT-NB0237": "Wolfgang Driza",
            "KTR-AT-NB0230": "Florian Fuchs",
            "KTR-AT-NB0228": "Martin Gaugitsch",
            "KTR-AT-NB0226": "Maryam Koulaei",
            "KTR-AT-NB0233": "Mirko Kosovic",
            "KTR-FR-NB0191": "Koffi Sidoine Alla",
            "KTR-FR-NB0019": "Jean-Pierre Brassier",
            "KTR-FR-NB0147": "Stéphane Duvois",
            "KTR-FR-NB0011": "Herve Le Guern",
            "KTR-FR-NB0123": "Jean-Pierre Recule",
            "KTR-FR-NB0012": "Philippe Potier",
            "KTR-FR-NB0149": "Kien Tendjoukian"
        }

        # Convertir la colonne des dates et heures en datetime
        try:
            df["Anti-Malware Updated On"] = pd.to_datetime(df["Anti-Malware Updated On"], format="%d/%m/%Y %H:%M", errors='coerce')
        except:
            messagebox.showerror("Erreur", "La colonne 'Anti-Malware Updated On' n'est pas au format de date et heure (xx/xx/xxxx HH:MM).")
            return

        # Filtrer les données en fonction de la liste des valeurs de "Computer Name" spécifiée
        computer_name_list = [
            "KTR-FR-NB0037", "KTR-FR-NB0090", "KTR-FR-NB0093", "KTR-FR-NB0094", "KTR-FR-NB0086",
            "KTR-FR-NB0040", "KTR-FR-NB0192", "KTR-AT-NB0155", "KTR-AT-NB0153", "KTR-AT-NB0157",
            "KTR-FR-NB0116", "KTR-FR-NB0065", "KTR-FR-NB0110", "KTR-FR-NB0124", "KTR-FR-NB0100",
            "KTR-FR-NB0099", "KTR-FR-NB0056", "KTR-FR-NB0115", "KTR-FR-NB0024", "KTR-FR-NB0140",
            "KTR-FR-NB0117", "KTR-FR-NB0097", "KTR-FR-NB0075", "KTR-FR-NB0154", "KTR-FR-NB0145",
            "KTR-FR-NB0190", "KTR-FR-NB0121", "KTR-FR-NB0092", "KTR-AT-NB0125", "KTR-DE-NB0029",
            "KTR-AT-NB0242", "KTR-AT-NB0209", "KTR-AT-NB0128", "KTR-AT-NB0068", "KTR-AT-NB0143",
            "KTR-AT-NB0139", "KTR-AT-NB0223", "KTR-AT-NB0249", "KTR-AT-NB0164", "KTR-AT-NB0064",
            "KTR-AT-NB0219", "KTR-AT-NB0220", "KTR-AT-NB0326", "KTR-AT-NB0135", "KTR-AT-NB0238",
            "KTR-AT-NB0318", "KTR-AT-NB0319", "KTR-PT-NB0209", "KTR-PT-NB0041", "KTR-PT-NB0213",
            "KTR-PT-NB0212", "KTR-PT-NB0206", "KTR-PT-NB0208", "KTR-PT-NB0211", "KTR-PT-NB0201",
            "KTR-PT-NB0210", "KTR-PT-NB0214", "KTR-FR-NB0028", "KTR-FR-NB0029", "KTR-FR-NB0135",
            "KTR-FR-NB0031", "KTR-FR-NB0180", "KTR-FR-NB0035", "KTR-FR-NB0157", "KTR-FR-NB0146",
            "KTR-FR-NB0130", "KTR-FR-NB0074", "KTR-FR-NB0041", "KTR-FR-NB0104", "KTR-FR-NB0128",
            "KTR-FR-NB0060", "KTR-FR-NB0073", "KTR-FR-NB0166", "KTR-FR-NB0106", "KTR-FR-NB0064",
            "KTR-FR-NB0061", "KTR-FR-NB0022", "KTR-AT-NB0245", "KTR-FR-NB0127", "KTR-FR-NB0047",
            "KTR-FR-NB0132", "KTR-FR-NB0138", "KTR-FR-NB0101", "KTR-AT-NB0049", "KTR-AT-NB0312",
            "KTR-FR-NB0174", "KTR-DE-NB0021", "KTR-AT-NB0241", "KTR-AT-NB0045", "KTR-BE-NB0093",
            "KTR-DE-NB0009", "KTR-UK-NB0020", "KTR-DE-NB0040", "KTR-AT-NB0156", "KTR-AT-NB0237",
            "KTR-AT-NB0230", "KTR-AT-NB0228", "KTR-AT-NB0226", "KTR-AT-NB0233", "KTR-FR-NB0191",
            "KTR-FR-NB0019", "KTR-FR-NB0147", "KTR-FR-NB0011", "KTR-FR-NB0123", "KTR-FR-NB0012",
            "KTR-FR-NB0149", "KTR-DE-NB0042"
        ]

        # Filtrer les données en fonction de la liste des valeurs de "Computer Name"
        df_filtered = df[df["Computer Name"].isin(computer_name_list)].copy()

        # Remplacer les valeurs de "Computer Name" par les noms complets correspondants
        df_filtered["Computer Name"] = df_filtered["Computer Name"].map(computer_name_associations)
        df_filtered["Nom et Prénom"] = df_filtered["Computer Name"].map(computer_name_associations)
        df_filtered["Date de mise à jour"] = df_filtered["Anti-Malware Updated On"].dt.date

        # Demander la date de création du fichier Excel
        date_jour = date_creation

        df_filtered.loc[:, "Durée depuis la mise à jour"] = date_jour - df_filtered["Date de mise à jour"]

        # Calculer les durées depuis la mise à jour
        delta_48h = timedelta(hours=48)
        delta_96h = timedelta(hours=96)
        delta_168h = timedelta(hours=168)

        

        # Filtrer les personnes en fonction de la durée depuis la mise à jour
        moins_de_48h = df_filtered[df_filtered["Durée depuis la mise à jour"] <= delta_48h]
        moins_de_96h = df_filtered[(df_filtered["Durée depuis la mise à jour"] > delta_48h) & (df_filtered["Durée depuis la mise à jour"] <= delta_96h)]
        moins_de_168h = df_filtered[(df_filtered["Durée depuis la mise à jour"] > delta_96h) & (df_filtered["Durée depuis la mise à jour"] <= delta_168h)]
        plus_de_168h = df_filtered[df_filtered["Durée depuis la mise à jour"] > delta_168h]
        
        # Afficher les résultats
        print(df_filtered)

        # Nombre de personnes dans chaque catégorie
        nb_personnes_moins_de_48h = len(moins_de_48h)
        nb_personnes_moins_de_96h = len(moins_de_96h)
        nb_personnes_moins_de_168h = len(moins_de_168h) 
        nb_personnes_plus_de_168h = len(plus_de_168h)

        # Pourcentage de personnes dans chaque catégorie
        total_personnes = len(df_filtered)
        pourcentage_moins_de_48h = (nb_personnes_moins_de_48h / total_personnes) * 100
        pourcentage_moins_de_96h = (nb_personnes_moins_de_96h / total_personnes) * 100
        pourcentage_moins_de_168h = (nb_personnes_moins_de_168h / total_personnes) * 100
        pourcentage_plus_de_168h = (nb_personnes_plus_de_168h / total_personnes) * 100

        # Liste de personnes dans chaque catégorie
        if "Computer Name" in moins_de_48h.columns:
            personnes_moins_de_48h = ", ".join(moins_de_48h["Computer Name"].tolist())
        else:
            personnes_moins_de_48h = ""

        if "Computer Name" in moins_de_96h.columns:
            personnes_moins_de_96h = ", ".join(moins_de_96h["Computer Name"].tolist())
        else:
            personnes_moins_de_96h = ""

        if "Computer Name" in moins_de_168h.columns:
            personnes_moins_de_168h = ", ".join(moins_de_168h["Computer Name"].tolist())
        else:
            personnes_moins_de_168h = ""

        if "Computer Name" in plus_de_168h.columns:
            personnes_plus_de_168h = ", ".join(plus_de_168h["Computer Name"].tolist())
        else:
            personnes_plus_de_168h = ""

        # Création d'un DataFrame avec les données structurées
        data = {
            "MAJ Malware -48h": [
                nb_personnes_moins_de_48h, pourcentage_moins_de_48h, personnes_moins_de_48h
            ],
            "MAJ Malware -96h": [
                nb_personnes_moins_de_96h, pourcentage_moins_de_96h, personnes_moins_de_96h
            ],
            "MAJ Malware -168h": [
                nb_personnes_moins_de_168h, pourcentage_moins_de_168h, personnes_moins_de_168h
            ],
            "MAJ Malware +168h": [
                nb_personnes_plus_de_168h, pourcentage_plus_de_168h, personnes_plus_de_168h
            ]
        }

        df_result = pd.DataFrame(data, index=["Nombre de personnes", "Pourcentage de personnes", "Liste des personnes"])

        # Afficher le DataFrame résultat
        print(df_result)

        # Création d'un DataFrame avec les données structurées
        data = {
            "MAJ Malware -48h": [
                nb_personnes_moins_de_48h, pourcentage_moins_de_48h, personnes_moins_de_48h
            ],
            "MAJ Malware -96h": [
                nb_personnes_moins_de_96h, pourcentage_moins_de_96h, personnes_moins_de_96h
            ],
            "MAJ Malware -168h": [
                nb_personnes_moins_de_168h, pourcentage_moins_de_168h, personnes_moins_de_168h
            ],
            "MAJ Malware +168h": [
                nb_personnes_plus_de_168h, pourcentage_plus_de_168h, personnes_plus_de_168h
            ]
        }

        df_resultat = pd.DataFrame(data, index=["Nombre de personnes", "Pourcentage", "Liste des personnes"])

        # Exporter les résultats vers un fichier Excel
        output_filename = "resultatsMajMalware.xlsx"  # Nom du fichier de sortie
        df_resultat.to_excel(output_filename, index_label="Catégorie")
        print(f"Les résultats ont été exportés vers le fichier '{output_filename}'.")

        messagebox.showinfo("Export terminé", f"Les résultats ont été exportés vers le fichier '{output_filename}'.")

    def select_file2(self, event):
        self.selected_file2 = filedialog.askopenfilename(initialdir="/", title="Select file", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
        self.drop_zone2.config(text=self.selected_file2)


    def drag_file2(self, event):
        self.drop_zone.config(bg="#32cf8e")
        self.dragging = True

    def release_file2(self, event):
        self.drop_zone.config(bg="#ffffff")
        self.dragging = False

    def on_enter2(self, event):
        if self.dragging:
            self.drop_zone.config(bg="#32cf8e")

    def on_leave2(self, event):
        if self.dragging:
            self.drop_zone.config(bg="#ffffff")

    def analyser_checkpoint_fichier(self):
        #  code pour analyser le fichier de checkpoint
        pass

    def analyser_checkpoint_fichier(self):
        if self.selected_file2:
            # Charger le fichier Excel avec pandas
            df = pd.read_excel(self.selected_file2)
            df = df[df['Computer Name'].isin(["KTR-FR-NB0037", "KTR-FR-NB0090", "KTR-FR-NB0093", "KTR-FR-NB0094", "KTR-FR-NB0086",
            "KTR-FR-NB0040", "KTR-FR-NB0192", "KTR-AT-NB0155", "KTR-AT-NB0153", "KTR-AT-NB0157",
            "KTR-FR-NB0116", "KTR-FR-NB0065", "KTR-FR-NB0110", "KTR-FR-NB0124", "KTR-FR-NB0100",
            "KTR-FR-NB0099", "KTR-FR-NB0056", "KTR-FR-NB0115", "KTR-FR-NB0024", "KTR-FR-NB0140",
            "KTR-FR-NB0117", "KTR-FR-NB0097", "KTR-FR-NB0075", "KTR-FR-NB0154", "KTR-FR-NB0145",
            "KTR-FR-NB0190", "KTR-FR-NB0121", "KTR-FR-NB0092", "KTR-AT-NB0125", "KTR-DE-NB0029",
            "KTR-AT-NB0242", "KTR-AT-NB0209", "KTR-AT-NB0128", "KTR-AT-NB0068", "KTR-AT-NB0143",
            "KTR-AT-NB0139", "KTR-AT-NB0223", "KTR-AT-NB0249", "KTR-AT-NB0164", "KTR-AT-NB0064",
            "KTR-AT-NB0219", "KTR-AT-NB0220", "KTR-AT-NB0326", "KTR-AT-NB0135", "KTR-AT-NB0238",
            "KTR-AT-NB0318", "KTR-AT-NB0319", "KTR-PT-NB0209", "KTR-PT-NB0041", "KTR-PT-NB0213",
            "KTR-PT-NB0212", "KTR-PT-NB0206", "KTR-PT-NB0208", "KTR-PT-NB0211", "KTR-PT-NB0201",
            "KTR-PT-NB0210", "KTR-PT-NB0214", "KTR-FR-NB0028", "KTR-FR-NB0029", "KTR-FR-NB0135",
            "KTR-FR-NB0031", "KTR-FR-NB0180", "KTR-FR-NB0035", "KTR-FR-NB0157", "KTR-FR-NB0146",
            "KTR-FR-NB0130", "KTR-FR-NB0074", "KTR-FR-NB0041", "KTR-FR-NB0104", "KTR-FR-NB0128",
            "KTR-FR-NB0060", "KTR-FR-NB0073", "KTR-FR-NB0166", "KTR-FR-NB0106", "KTR-FR-NB0064",
            "KTR-FR-NB0061", "KTR-FR-NB0022", "KTR-AT-NB0245", "KTR-FR-NB0127", "KTR-FR-NB0047",
            "KTR-FR-NB0132", "KTR-FR-NB0138", "KTR-FR-NB0101", "KTR-AT-NB0049", "KTR-AT-NB0312",
            "KTR-FR-NB0174", "KTR-DE-NB0021", "KTR-AT-NB0241", "KTR-AT-NB0045", "KTR-BE-NB0093",
            "KTR-DE-NB0009", "KTR-UK-NB0020", "KTR-DE-NB0040", "KTR-AT-NB0156", "KTR-AT-NB0237",
            "KTR-AT-NB0230", "KTR-AT-NB0228", "KTR-AT-NB0226", "KTR-AT-NB0233", "KTR-FR-NB0191",
            "KTR-FR-NB0019", "KTR-FR-NB0147", "KTR-FR-NB0011", "KTR-FR-NB0123", "KTR-FR-NB0012",
            "KTR-FR-NB0149", "KTR-DE-NB0042"])]

            # Correspondances entre noms d'ordinateur et noms et prénoms
            correspondances = {
            "KTR-FR-NB0037": "Bobi Andrieski",
            "KTR-FR-NB0090": "Daniel Bouilly",
            "KTR-FR-NB0093": "Philippe Dardenne",
            "KTR-FR-NB0094": "Christophe Leroux",
            "KTR-FR-NB0086": "Fernand Mendes",
            "KTR-FR-NB0040": "Cedric Valynseele",
            "KTR-FR-NB0192": "Youssef Tabaddine",
            "KTR-AT-NB0155": "Martin Hollmuller",
            "KTR-AT-NB0153": "Gunther Lempert",
            "KTR-AT-NB0157": "Max Linhart",
            "KTR-FR-NB0116": "Jalel Bouachir",
            "KTR-FR-NB0065": "Victor Cardoso",
            "KTR-FR-NB0110": "Olivier Daunas",
            "KTR-FR-NB0124": "Nicolas De Bouvier",
            "KTR-FR-NB0100": "Driss Erouah",
            "KTR-FR-NB0099": "Jean Paul Fourdrinois",
            "KTR-FR-NB0056": "Pascal Garimede",
            "KTR-FR-NB0115": "Olivier Gouzerh",
            "KTR-FR-NB0024": "Samir Hamiane",
            "KTR-FR-NB0140": "Frederic Jacotot",
            "KTR-FR-NB0117": "Frederic Jan",
            "KTR-FR-NB0097": "Dominique Nabais",
            "KTR-FR-NB0075": "Christine Pietri",
            "KTR-FR-NB0154": "Tsiry Lalaina Ramanantsialonina",
            "KTR-FR-NB0145": "Christophe Surdive",
            "KTR-FR-NB0190": "Olivier Taieb",
            "KTR-FR-NB0121": "Thierry Vincent",
            "KTR-FR-NB0092": "Redouane Dahmani",
            "KTR-AT-NB0125": "Gerald Balham",
            "KTR-DE-NB0029": "Quang-Hai Dao",
            "KTR-AT-NB0242": "György Hepka",
            "KTR-AT-NB0209": "Hüseyin Kahraman",
            "KTR-AT-NB0128": "Robert Klevar",
            "KTR-AT-NB0068": "Thomas Linsberger",
            "KTR-AT-NB0143": "Michael Neumann",
            "KTR-AT-NB0139": "Rezart Para",
            "KTR-AT-NB0223": "Helmut Scheiber",
            "KTR-AT-NB0249": "Hermann Schwarzinger",
            "KTR-AT-NB0164": "Christian Hans Stark",
            "KTR-AT-NB0064": "Christoph Strasser",
            "KTR-AT-NB0219": "Dominik Szanva",
            "KTR-AT-NB0220": "Jeremias Szlezak",
            "KTR-AT-NB0326": "Daniel Urabl",
            "KTR-AT-NB0135": "Eduard Wittmann",
            "KTR-AT-NB0238": "Vladislav Zinovenko",
            "KTR-AT-NB0318": "Luca Zinicola",
            "KTR-AT-NB0319": "Omar Zubair",
            "KTR-PT-NB0209": "Jorge Caetano",
            "KTR-PT-NB0041": "Elizabeth Fernandez",
            "KTR-PT-NB0213": "Ricardo Monteiro",
            "KTR-PT-NB0212": "Bruno Oliveira",
            "KTR-PT-NB0206": "Filipe Soares Pereira",
            "KTR-PT-NB0208": "Filipe Pires",
            "KTR-PT-NB0211": "Miguel Ramos",
            "KTR-PT-NB0201": "Susana Rocha",
            "KTR-PT-NB0210": "Nuno Santos",
            "KTR-PT-NB0214": "Igor Videra Soares",
            "KTR-FR-NB0028": "Kaci Anis Derridi",
            "KTR-FR-NB0029": "Hélène Farine",
            "KTR-FR-NB0135": "Mickael Lambert",
            "KTR-FR-NB0031": "Ludovic Martin",
            "KTR-FR-NB0180": "Mamadou Seck",
            "KTR-FR-NB0035": "Frederic Pesson",
            "KTR-FR-NB0157": "Thibault Rabe",
            "KTR-FR-NB0146": "Elhadi Thierno Sadou Diallo",
            "KTR-FR-NB0130": "Sylia Baraka",
            "KTR-FR-NB0074": "Alain Caille",
            "KTR-FR-NB0041": "Claudy Cavallari",
            "KTR-FR-NB0104": "Jean-Tycho Dollfus",
            "KTR-FR-NB0128": "Laurent Dorys-Charnalet",
            "KTR-FR-NB0060": "David Gastao",
            "KTR-FR-NB0073": "Trung Lim Cao",
            "KTR-FR-NB0166": "Antoine Jegoux",
            "KTR-FR-NB0106": "Olivier Marty",
            "KTR-FR-NB0064": "Benjamina Rabarison",
            "KTR-FR-NB0061": "Florent Roblin",
            "KTR-FR-NB0022": "David Roux",
            "KTR-AT-NB0245": "Levente Csakvari",
            "KTR-FR-NB0127": "Thierry Demarche",
            "KTR-FR-NB0047": "Fatiha Devaud",
            "KTR-FR-NB0132": "Lony Fischer",
            "KTR-FR-NB0138": "Laroussi Gattoufi",
            "KTR-FR-NB0101": "Yann Le Cor",
            "KTR-AT-NB0049": "Andreas Mayrhofer",
            "KTR-AT-NB0312": "Dominik Neznik",
            "KTR-FR-NB0174": "Reva Rabenandrasana",
            "KTR-DE-NB0021": "Ole Hohlt",
            "KTR-DE-NB0042": "Ndubueze Ogaluzo",
            "KTR-AT-NB0241": "Atanas Pechevski",
            "KTR-AT-NB0045": "Martin Riegler",
            "KTR-BE-NB0093": "Endri Stefani",
            "KTR-DE-NB0009": "Erich Seitz",
            "KTR-UK-NB0020": "Therese Taylor",
            "KTR-DE-NB0040": "Onur Ulu",
            "KTR-AT-NB0156": "Gunter Bauer",
            "KTR-AT-NB0237": "Wolfgang Driza",
            "KTR-AT-NB0230": "Florian Fuchs",
            "KTR-AT-NB0228": "Martin Gaugitsch",
            "KTR-AT-NB0226": "Maryam Koulaei",
            "KTR-AT-NB0233": "Mirko Kosovic",
            "KTR-FR-NB0191": "Koffi Sidoine Alla",
            "KTR-FR-NB0019": "Jean-Pierre Brassier",
            "KTR-FR-NB0147": "Stéphane Duvois",
            "KTR-FR-NB0011": "Herve Le Guern",
            "KTR-FR-NB0123": "Jean-Pierre Recule",
            "KTR-FR-NB0012": "Philippe Potier",
            "KTR-FR-NB0149": "Kien Tendjoukian"
            }

            # Ajouter la colonne "Nom et Prénom"
            df['Nom et Prénom'] = df['Computer Name'].map(correspondances)

            # Calculer le nombre de personnes ayant la dernière version de Endpoint Version
            derniere_version = df['Endpoint Version'].max()
            derniere_version_count = df[df['Endpoint Version'] == derniere_version].shape[0]
            derniere_version_computer_names = df[df['Endpoint Version'] == derniere_version]['Computer Name'].tolist()

            # Ajouter la ligne pour remplacer les noms d'ordinateur par les noms et prénoms correspondants
            df['Computer Name'] = df['Computer Name'].map(correspondances)

            total_personnes = df.shape[0]
            pourcentage_derniere_version = (derniere_version_count / total_personnes) * 100

            # Filtrer les personnes avec la version N-1 de Endpoint Version
            versions_uniques = df['Endpoint Version'].unique()
            versions_uniques_sorted = sorted(versions_uniques, reverse=True)
            if len(versions_uniques_sorted) >= 2:
                version_N_1 = versions_uniques_sorted[1]
                version_N_1_count = df[df['Endpoint Version'] == version_N_1].shape[0]
                version_N_1_pourcentage = (version_N_1_count / total_personnes) * 100
                version_N_1_computer_names = df[df['Endpoint Version'] == version_N_1]['Computer Name'].tolist()
            else:
                version_N_1_count = 0
                version_N_1_pourcentage = 0
                version_N_1_computer_names = []

            # Filtrer les personnes avec la version N-2 de Endpoint Version
            if len(versions_uniques_sorted) >= 3:
                version_N_2 = versions_uniques_sorted[2]
                version_N_2_count = df[df['Endpoint Version'] == version_N_2].shape[0]
                version_N_2_pourcentage = (version_N_2_count / total_personnes) * 100
                version_N_2_computer_names = df[df['Endpoint Version'] == version_N_2]['Computer Name'].tolist()
            else:
                version_N_2_count = 0
                version_N_2_pourcentage = 0
                version_N_2_computer_names = []

            # Création du DataFrame pour l'export vers Excel
            data = {
                "Dernière Version": [
                    derniere_version_count, pourcentage_derniere_version, ", ".join(derniere_version_computer_names)
                ],
                "Version N-1": [
                    version_N_1_count, version_N_1_pourcentage, ", ".join(version_N_1_computer_names)
                ],
                "Version N-2": [
                    version_N_2_count, version_N_2_pourcentage, ", ".join(version_N_2_computer_names)
                ],
            }

            df_resultat = pd.DataFrame(data, index=["Nombre de personnes", "Pourcentage", "Liste des ordinateurs"])

            # Exporter le DataFrame vers un fichier Excel
            df_resultat.to_excel("resultatVersionEndpoint.xlsx")

            messagebox.showinfo("Résultats", "Les résultats ont été exportés vers le fichier 'resultatVersionEndpoint.xlsx'.")
        else:
            messagebox.showwarning("Erreur", "Aucun fichier Excel sélectionné (corp2)")
    def deconnecter_application(self):
        result = messagebox.askquestion("Déconnexion", "Êtes-vous sûr de vouloir vous déconnecter ?", icon="warning")
        if result == "yes":
            self.root.destroy()

    def quitter_application(self):
        result = messagebox.askquestion("Quitter", "Êtes-vous sûr de vouloir quitter l'application ?", icon="warning")
        if result == "yes":
            sys.exit()

root = Tk()
obj = Dashboard(root)
root.mainloop()
# fin du code        
